import pandas as pd
import os
import openpyxl
import datetime
import locale
from prev import previsao

pd.options.display.float_format = "{:,.2f}".format
locale.setlocale(locale.LC_TIME, "pt_BR")

# juntar todos os arquivos em um unico data frame
new_df = pd.DataFrame()
for file in sorted(os.listdir("dataset/files"), key=lambda x: x[7]):
    df = pd.read_excel(f"dataset/files/{file}")
    new_df = pd.concat([new_df, df], ignore_index=True)

# criar wb para os relatorios e a pagina do relatorio geral
wb = openpyxl.Workbook()
wb.create_sheet("Relatório Geral")
del wb["Sheet"]

ws = wb["Relatório Geral"]

# data frame do relatorio geral
data = {
    "Período": [],
    "Vendas Totais (R$)": [],
    "Produtos Vendidos": [],
    "Categoria Destaque (Maior quantidade vendida)": [],
    "Região Destaque (Mais rendimento)": [],
    "Previsão Próximo Mês (R$)": []
}

# adicionar o periodo
meses = new_df["Data"].dt.strftime("%B").values.tolist()
for mes in meses:
    if f"{mes}/{datetime.datetime.today().year}" not in data["Período"]:
        periodo = f"{mes}/{datetime.datetime.today().year}"
        data["Período"].append(periodo)



# Vendas totais por mes e produtos vendidos
df_rel = new_df.groupby(new_df["Data"].dt.strftime("%B")).aggregate({"Preço Total (R$)": "sum", "Quantidade": "sum"})

# para cada periodo, percorrendo qual o valor que esta em df_rel e adicionando na "linha" certa do data
for periodo in data["Período"]:
    periodo = periodo.split("/")[0]
    
    for index, row in df_rel.iterrows():
        venda_total = float(row.values[0])
        quant_total = int(row.values[1])
        if index == periodo:
            data["Vendas Totais (R$)"].append(venda_total)
            data["Produtos Vendidos"].append(quant_total)
            
            
            
# agrupar por categoria
df_rel = new_df.groupby([new_df["Data"].dt.strftime("%B"), "Categoria"])["Quantidade"].sum().to_frame()
for periodo in data["Período"]:
    periodo = periodo.split("/")[0]
    categoria = df_rel.loc[periodo].idxmax().values[0]

    for index, row in df_rel.iterrows():
        if index[0] == periodo:
            data["Categoria Destaque (Maior quantidade vendida)"].append(categoria)
            break


# agrupar por regiao
df_rel = new_df.groupby([new_df["Data"].dt.strftime("%B"), "Região"])["Preço Total (R$)"].sum().to_frame()
for periodo in data["Período"]:
    periodo = periodo.split("/")[0]
    regiao = df_rel.loc[periodo].idxmax().values[0]
    
    for index, row in df_rel.iterrows():
        if index[0] == periodo:
            data["Região Destaque (Mais rendimento)"].append(regiao)
            break



# PREVISAO (MACHINE LEARNING)

df = pd.DataFrame({
    "Período": data["Período"],
    "Quantidade Vendida": data["Produtos Vendidos"],
    "Vendas Totais": data["Vendas Totais (R$)"]
})

df_prev = previsao(df)
for index, row in df_prev.iterrows():
    periodo_prev = row.iloc[0]
    prev_value = round(row.iloc[7], 2)
    
    for i in range(len(data["Período"]) - 1):
        try:
            ind = data["Período"].index(periodo_prev)
            data["Previsão Próximo Mês (R$)"][ind] = prev_value
        except IndexError:
            data["Previsão Próximo Mês (R$)"].append("Dados Insuficientes")



df = pd.DataFrame(data).to_excel("output/files/Relatório_Geral.xlsx", sheet_name="Relatório Geral", index=False)

wb = openpyxl.load_workbook("output/files/Relatório_Geral.xlsx")
ws = wb["Relatório Geral"]
    
for row in ws.iter_rows(min_row=2):
    vendas = row[1]
    prev_vendas = row[5]
        
    vendas.value = round(vendas.value, 2)
    if prev_vendas.value != "Dados Insuficientes":
        prev_vendas.value = round(prev_vendas.value, 2)
        
    vendas.number_format = 'R$ #,##0.00'
    prev_vendas.number_format = 'R$ #,##0.00'
wb.save("output/files/Relatório_Geral.xlsx")